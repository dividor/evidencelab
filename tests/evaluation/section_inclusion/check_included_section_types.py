#!/usr/bin/env python3
"""
Check that every section inside a document's main-body page range is tagged with
a section type that is *included* by default in Search / Content settings.

Background
----------
Search has a "content type" filter. By default it only returns chunks whose
``tag_section_type`` is one of the "real content" types
(``DEFAULT_INCLUDED_SECTION_TYPES`` below, mirrored from the frontend constant
``DEFAULT_SECTION_TYPES`` in ``ui/frontend/src/utils/searchUrl.ts``). Any other
section type (front_matter, acronyms, annexes, appendix, bibliography,
introduction, ...) is *excluded* by default.

Section tagging is automatic (done at upload by the TOC classifier), so a body
section that is mis-tagged as, say, ``annexes`` silently disappears from default
search results. Each document also carries a human-set page range in its source
metadata:

    "Introduction - before beginning of Annexes (start_page_number, end_page_number)"

That range is the ground-truth for "where the real body content lives". This
script flags every classified-TOC section whose page falls inside that range but
whose section type is NOT in the default-included set — i.e. body content that
would be accidentally excluded from default search.

Data sources (all from the Postgres sidecar, ``docs_<data_source>``):
  * body page range -> ``src_doc_raw_metadata[INTRO_RANGE_FIELD]``, e.g. "(11, 62)"
  * section tags    -> ``sys_data["sys_toc_classified"]``, the same tags shown in
                       the "Contents" tab of the document viewer.

Usage
-----
    # Check all documents in a data source
    python tests/evaluation/section_inclusion/check_included_section_types.py --data-source wfp

    # Check a sample of N documents
    python tests/evaluation/section_inclusion/check_included_section_types.py --records 20

    # Check a single document by doc id
    python tests/evaluation/section_inclusion/check_included_section_types.py --file-id <id>

    # Custom output path
    python tests/evaluation/section_inclusion/check_included_section_types.py --output results.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))  # noqa: E402

# Mirrors DEFAULT_SECTION_TYPES in ui/frontend/src/utils/searchUrl.ts — the
# section types that Search includes by default. Everything else is excluded.
DEFAULT_INCLUDED_SECTION_TYPES = (
    "executive_summary",
    "context",
    "methodology",
    "findings",
    "conclusions",
    "recommendations",
    "other",
)

# Source-metadata field holding the human-set body page range, e.g. "(11, 62)".
INTRO_RANGE_FIELD = (
    "Introduction - before beginning of Annexes (start_page_number, end_page_number)"
)

# Only look at documents that finished indexing.
ALLOWED_STATUSES = {"indexed", "tagged"}

# Matches classified-TOC lines, e.g.:
#   "[H1] 1. Introduction | introduction | page 11"
#   "[H2] Table des Matières | front_matter | page 3 (i) [Front]"
# The trailing bracket marker ([Front], [FM], ...) and the roman-numeral page
# alias are both optional.
TOC_CLASSIFIED_PATTERN = re.compile(
    r"^\s*\[H(?P<level>\d+)\]\s*(?P<title>.*?)\s*\|\s*(?P<label>[^|]+?)"
    r"(?:\s*\|\s*page\s*(?P<page>\d+)(?:\s*\([^)]+\))?)?"
    r"(?:\s*\[[^\]]*\])?\s*$",
    flags=re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Flag sections in the main-body page range that are tagged with a "
            "section type excluded from Search by default."
        ),
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--records",
        type=int,
        default=None,
        help="Number of documents to process (default: all)",
    )
    group.add_argument(
        "--file-id",
        type=str,
        help="Process a single document by doc id",
    )
    parser.add_argument(
        "--data-source",
        type=str,
        default="wfp",
        help="Data source / table suffix (default: wfp)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=str(PROJECT_ROOT / "logs" / "section_inclusion_eval.xlsx"),
        help="Output Excel (.xlsx) file path",
    )
    return parser.parse_args()


# --------------------------------------------------------------------------- #
# Pure logic (unit-tested in tests/unit/test_check_included_section_types.py)
# --------------------------------------------------------------------------- #
def _parse_page_value(value: Any) -> Optional[int]:
    """Coerce a single page value (int/float/numeric-string) to int."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def parse_page_range(raw_value: Any) -> Tuple[Optional[int], Optional[int]]:
    """Parse the body page range from the shapes it can take.

    In practice the stored value is a string like ``"(11, 62)"``, but tuples,
    lists, dicts and bare ints are handled too. Returns ``(start, end)``; either
    element may be ``None`` if it could not be determined.
    """
    if raw_value is None:
        return None, None

    if isinstance(raw_value, (list, tuple)) and len(raw_value) >= 2:
        return _parse_page_value(raw_value[0]), _parse_page_value(raw_value[1])

    if isinstance(raw_value, dict):
        start = raw_value.get("start_page_number", raw_value.get("start"))
        end = raw_value.get("end_page_number", raw_value.get("end"))
        return _parse_page_value(start), _parse_page_value(end)

    if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
        page = _parse_page_value(raw_value)
        return page, page

    if isinstance(raw_value, str):
        numbers = [int(val) for val in re.findall(r"\d+", raw_value)]
        if len(numbers) >= 2:
            return numbers[0], numbers[1]
        if len(numbers) == 1:
            return numbers[0], numbers[0]

    return None, None


def parse_toc_classified(toc_text: str) -> List[Dict[str, Any]]:
    """Parse classified-TOC text into ``{title, label, page}`` entries."""
    if not toc_text:
        return []

    entries: List[Dict[str, Any]] = []
    for raw_line in toc_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = TOC_CLASSIFIED_PATTERN.match(line)
        if not match:
            continue
        page_text = match.group("page")
        entries.append(
            {
                "title": (match.group("title") or "").strip(),
                "label": (match.group("label") or "").strip().lower(),
                "page": int(page_text) if page_text and page_text.isdigit() else None,
            }
        )
    return entries


def find_excluded_sections(
    entries: Iterable[Dict[str, Any]],
    start_page: int,
    end_page: int,
    included: Iterable[str] = DEFAULT_INCLUDED_SECTION_TYPES,
) -> List[Dict[str, Any]]:
    """Return TOC entries whose page is in ``[start_page, end_page]`` and whose
    section type would be excluded from Search by default.

    A section is a violation when its label is not in ``included`` — that content
    lives in the main body but would be filtered out of default search results.
    """
    included_set = set(included)
    violations: List[Dict[str, Any]] = []
    for entry in entries:
        page = entry.get("page")
        if page is None or not (start_page <= page <= end_page):
            continue
        if entry.get("label") not in included_set:
            violations.append(entry)
    return violations


def _describe_sections(sections: List[Dict[str, Any]]) -> str:
    """Human-readable one-line summary of violating sections for the report."""
    parts = []
    for sec in sections:
        label = sec.get("label") or "(unlabeled)"
        parts.append(f"[{label}] {sec.get('title', '')} (p.{sec.get('page')})")
    return "; ".join(parts)


# --------------------------------------------------------------------------- #
# Document accessors (rows as returned by PostgresClient.fetch_all_docs())
# --------------------------------------------------------------------------- #
def get_document_title(doc: dict) -> str:
    raw = doc.get("src_doc_raw_metadata") or {}
    return (
        doc.get("map_title")
        or raw.get("Title evaluation")
        or raw.get("title")
        or "Unknown"
    )


def get_toc_classified(doc: dict) -> str:
    sys_data = doc.get("sys_data") or {}
    return sys_data.get("sys_toc_classified") or doc.get("sys_toc_classified") or ""


def get_intro_range_field(doc: dict) -> Any:
    raw = doc.get("src_doc_raw_metadata") or {}
    return raw.get(INTRO_RANGE_FIELD)


def _count_sections_in_range(
    entries: List[Dict[str, Any]], start_page: Optional[int], end_page: Optional[int]
) -> int:
    if start_page is None or end_page is None:
        return 0
    return sum(
        1
        for entry in entries
        if entry.get("page") is not None and start_page <= entry["page"] <= end_page
    )


def evaluate_document(doc: dict) -> Dict[str, Any]:
    """Evaluate a single document row, returning a flat result row."""
    intro_range_raw = get_intro_range_field(doc)
    start_page, end_page = parse_page_range(intro_range_raw)
    entries = parse_toc_classified(get_toc_classified(doc))

    reasons: List[str] = []
    if not entries:
        reasons.append("missing_toc_classified")
    if start_page is None or end_page is None:
        reasons.append("missing_metadata_range")

    violations: List[Dict[str, Any]] = []
    if not reasons:
        violations = find_excluded_sections(entries, start_page, end_page)

    if reasons:
        status = "skipped"
    elif violations:
        status = "fail"
    else:
        status = "pass"

    excluded_types = sorted({sec.get("label") or "(unlabeled)" for sec in violations})
    return {
        "title": get_document_title(doc),
        "metadata_range": str(intro_range_raw),
        "range_start": start_page,
        "range_end": end_page,
        "sections_in_range": _count_sections_in_range(entries, start_page, end_page),
        "num_excluded": len(violations),
        "excluded_section_types": ", ".join(excluded_types),
        "excluded_details": _describe_sections(violations),
        "status": status,
        "reasons": ", ".join(reasons),
    }


def select_documents(
    docs: List[Dict[str, Any]], limit: Optional[int], file_id: Optional[str]
) -> List[Dict[str, Any]]:
    """Filter fetched docs down to the ones this run should evaluate."""
    if file_id:
        selected = [doc for doc in docs if str(doc.get("id")) == str(file_id)]
        if not selected:
            raise ValueError(f"Document {file_id} not found")
        return selected

    eligible = [doc for doc in docs if doc.get("sys_status") in ALLOWED_STATUSES]
    return eligible[:limit] if limit else eligible


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #
REPORT_HEADERS = [
    "doc_id",
    "title",
    "data_source",
    "metadata_range",
    "range_start",
    "range_end",
    "sections_in_range",
    "num_excluded",
    "excluded_section_types",
    "excluded_details",
    "status",
    "reasons",
]


def _create_workbook():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Section Inclusion"
    ws.append(REPORT_HEADERS)
    header_fill = PatternFill(
        start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return wb, ws


def _print_summary(counts: Dict[str, int], excluded_tally: Dict[str, int]) -> None:
    total = sum(counts.values())
    print("\n" + "=" * 60)
    print(f"Documents processed : {total}")
    print(f"  pass    : {counts['pass']}")
    print(f"  fail    : {counts['fail']}  (body content excluded by default)")
    print(f"  skipped : {counts['skipped']}  (no range / no classified TOC)")
    if excluded_tally:
        print("\nExcluded section types found inside body ranges:")
        print("(documents affected, per section type)")
        for label, count in sorted(excluded_tally.items(), key=lambda kv: -kv[1]):
            print(f"  {label:<20} {count}")
    print("=" * 60)


def main() -> None:
    from pipeline.db import PostgresClient

    args = parse_args()
    pg = PostgresClient(args.data_source)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb, ws = _create_workbook()

    counts = {"pass": 0, "fail": 0, "skipped": 0}
    excluded_tally: Dict[str, int] = {}

    for doc in select_documents(pg.fetch_all_docs(), args.records, args.file_id):
        result = evaluate_document(doc)
        counts[result["status"]] += 1
        for label in result["excluded_section_types"].split(", "):
            if label:
                excluded_tally[label] = excluded_tally.get(label, 0) + 1
        ws.append(
            [str(doc.get("id")), result["title"], args.data_source]
            + [result[key] for key in REPORT_HEADERS[3:]]
        )
        if result["status"] == "fail":
            print(f"FAIL {str(doc.get('id'))[:8]}  {result['title'][:60]}")
            print(f"     {result['excluded_details'][:200]}")

    wb.save(str(output_path))
    _print_summary(counts, excluded_tally)
    print(f"\nReport written to: {output_path}")


if __name__ == "__main__":
    main()
