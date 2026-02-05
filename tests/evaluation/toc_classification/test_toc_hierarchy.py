#!/usr/bin/env python3
"""
Evaluate TOCs stored in the database using the LLM judge.

This script pulls TOCs from the DB (prefers toc_classified when available),
extracts raw TOC text from the PDF, and runs the LLM judge to score alignment.
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

import fitz  # noqa: E402
import openpyxl  # noqa: E402
from jinja2 import Environment, FileSystemLoader  # noqa: E402
from langchain_core.messages import HumanMessage  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

from pipeline.db import Database, load_datasources_config  # noqa: E402
from utils.llm_factory import get_llm  # noqa: E402

# Initialize Jinja2 environment for loading templates
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
PROMPTS_DIR = PROJECT_ROOT / "prompts"
jinja_env = Environment(loader=FileSystemLoader(str(PROMPTS_DIR)))

# Evaluation model configuration
JUDGE_LLM = "Qwen/Qwen2.5-72B-Instruct"
MAX_RENDERED_PROMPT_CHARS = 32000
MAX_PDF_TOC_CHARS = 12000
LOG_PATH = PROJECT_ROOT / "logs" / "toc_evaluation.log"


def parse_args():
    parser = argparse.ArgumentParser(description="Judge TOCs stored in the database")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--records", type=int, help="Number of records to process")
    group.add_argument(
        "--file-id", type=str, help="Process a single document by file ID"
    )
    parser.add_argument(
        "--output",
        type=str,
        default="toc_evaluation_results.xlsx",
        help="Output Excel file path",
    )
    parser.add_argument(
        "--data-source",
        type=str,
        default="uneg",
        help="Data source/collection suffix (default: uneg)",
    )
    return parser.parse_args()


def get_test_documents(limit: int, data_source: str = "uneg"):
    """Fetch documents that already have TOCs from the specified data source."""
    db = Database(data_source=data_source)
    collection_name = db.documents_collection

    print(f"Querying {collection_name} for documents with TOCs...")

    results = []
    next_offset = None
    while len(results) < limit:
        points, next_offset = db.client.scroll(
            collection_name=collection_name,
            limit=max(limit - len(results), 50),
            with_payload=True,
            offset=next_offset,
        )
        if not points:
            break
        for point in points:
            if _payload_has_toc(point.payload):
                results.append(point)
                if len(results) >= limit:
                    break
        if next_offset is None:
            break

    return results


def get_document_by_id(file_id: str):
    """Fetch a single document by file ID from any collection."""
    config = load_datasources_config()
    datasources = config.get("datasources", {})

    db = Database()
    client = db.client

    if isinstance(datasources, dict):
        for ds_name, ds_config in datasources.items():
            if not isinstance(ds_config, dict):
                continue

            subdir = ds_config.get("data_subdir")
            if not subdir:
                continue

            collection_name = f"documents_{subdir}"
            ids_to_try: list[str | int] = [file_id]
            if file_id.isdigit():
                ids_to_try.append(int(file_id))

            for id_val in ids_to_try:
                try:
                    points = client.retrieve(
                        collection_name=collection_name,
                        ids=[id_val],
                        with_payload=True,
                    )
                    if points:
                        print(f"Found document in {collection_name} ({ds_name})")
                        return points, subdir
                except Exception:
                    continue

    raise ValueError(f"Document {file_id} not found in any collection.")


def get_corrected_toc_from_payload(payload: dict) -> str:
    """Prefer toc_classified, then fallback to toc fields."""
    toc_classified = payload.get("toc_classified") or payload.get("sys_toc_classified")
    if toc_classified:
        return toc_classified

    toc_data = payload.get("toc") or payload.get("sys_toc") or ""
    if isinstance(toc_data, str):
        return toc_data
    if isinstance(toc_data, list):
        return _format_toc_list(toc_data)
    return ""


def get_document_title(payload: dict) -> str:
    return (
        payload.get("title")
        or payload.get("map_title")
        or payload.get("src_title")
        or payload.get("src_title_sanitized")
        or payload.get("src_Title evaluation")
        or payload.get("map_title")
        or "Unknown"
    )


def get_pdf_web_link(payload: dict) -> str:
    return (
        payload.get("pdf_url")
        or payload.get("src_pdf_url")
        or payload.get("src_pdf")
        or payload.get("pdf_link")
        or payload.get("url")
        or ""
    )


def extract_pdf_toc_for_validation(payload: dict) -> tuple[str, str]:
    """Extract raw PDF TOC text for validation."""
    pdf_path, path_error = _resolve_pdf_path(payload)
    if not pdf_path:
        return "", path_error

    toc_entries = parse_toc_entries(payload)
    page_count = payload.get("page_count") or payload.get("sys_page_count")
    start_page, end_page = detect_toc_pages_from_pdf_content(
        pdf_path, max_scan_pages=10, page_count=page_count
    )
    if not start_page:
        start_page, end_page = get_toc_page_range(toc_entries, total_pages=page_count)
    if not start_page:
        return "", "Unable to determine TOC page range from payload or PDF content"

    raw_toc_text = fetch_page_text_from_pdf(pdf_path, start_page, end_page)
    if not raw_toc_text:
        return "", f"No text extracted from PDF pages {start_page}-{end_page}"

    filtered_text = filter_table_figure_box_entries(raw_toc_text)
    eval_text = _extract_toc_lines(filtered_text, page_count)
    return eval_text or filtered_text, ""


def parse_toc_entries(payload: dict) -> List[Dict[str, Any]]:
    toc_data = payload.get("toc") or payload.get("sys_toc")
    if isinstance(toc_data, list):
        return [
            {
                "page": entry.get("page", -1),
                "level": entry.get("level", 1),
                "title": entry.get("label") or entry.get("title") or entry.get("text"),
            }
            for entry in toc_data
            if isinstance(entry, dict)
        ]
    if isinstance(toc_data, str):
        return _parse_toc_string(toc_data)
    return []


def _parse_toc_string(toc_text: str) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    for line in toc_text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(
            r"^\s*\[H(?P<level>\d+)\]\s*(?P<title>.*?)\s*\|\s*page\s*(?P<page>\S+)\s*$",
            line,
        )
        if not match:
            continue
        page_raw = match.group("page")
        page = int(page_raw) if page_raw.isdigit() else -1
        entries.append(
            {
                "page": page if page > 0 else -1,
                "level": int(match.group("level")),
                "title": match.group("title").strip(),
            }
        )
    return entries


def get_toc_page_range(
    toc_entries: Iterable[Dict[str, Any]],
    total_pages: Optional[int] = None,
) -> Tuple[Optional[int], Optional[int]]:
    pages = [entry["page"] for entry in toc_entries if entry.get("page", -1) > 0]
    if not pages:
        return None, None
    start_page = min(pages)
    end_page = max(pages)
    if total_pages is not None and end_page > total_pages:
        end_page = total_pages
    return start_page, end_page


def detect_toc_pages_from_pdf_content(
    pdf_path: str, max_scan_pages: int = 10, page_count: Optional[int] = None
) -> Tuple[Optional[int], Optional[int]]:
    keywords = [
        "table of contents",
        "contents",
        "sommaire",
        "indice",
        "tabla de contenido",
        "tabla de contenidos",
        "index",
    ]
    start_page: Optional[int] = None
    end_page: Optional[int] = None

    doc = fitz.open(pdf_path)
    total_pages = doc.page_count
    scan_pages = min(max_scan_pages, total_pages)

    toc_like_counts = []
    toc_keyword_pages = []

    try:
        for page_index in range(scan_pages):
            text = doc.load_page(page_index).get_text("text")
            lower = text.lower()
            if any(keyword in lower for keyword in keywords):
                toc_keyword_pages.append(page_index)
            toc_like_counts.append(_count_toc_like_lines(text, page_count=page_count))

        if toc_keyword_pages:
            start_page = toc_keyword_pages[0] + 1
            end_page = start_page
            for next_index in range(toc_keyword_pages[0] + 1, scan_pages):
                if toc_like_counts[next_index] >= 5:
                    end_page = next_index + 1
                else:
                    break
            return start_page, end_page

        best_index = max(
            range(scan_pages), key=lambda i: toc_like_counts[i], default=-1
        )
        if best_index >= 0 and toc_like_counts[best_index] >= 5:
            start_page = best_index + 1
            end_page = start_page
            for next_index in range(best_index + 1, scan_pages):
                if toc_like_counts[next_index] >= 5:
                    end_page = next_index + 1
                else:
                    break
    finally:
        doc.close()

    return start_page, end_page


def fetch_page_text_from_pdf(pdf_path: str, start_page: int, end_page: int) -> str:
    if start_page < 1 or end_page < start_page:
        raise ValueError("Invalid page range")

    doc = fitz.open(pdf_path)
    try:
        lines: List[str] = []
        for page_index in range(start_page - 1, end_page):
            text = doc.load_page(page_index).get_text("text")
            lines.append(text)
        return "\n".join(lines).strip()
    finally:
        doc.close()


def filter_table_figure_box_entries(text: str) -> str:
    if not text:
        return ""

    header_patterns = [
        r"^list of figures$",
        r"^list of tables$",
        r"^list of boxes$",
        r"^lista de figuras$",
        r"^lista de tablas$",
        r"^liste des figures$",
        r"^liste des tableaux$",
        r"^lista de cuadros$",
    ]
    entry_patterns = [
        r"^(figure|figura|fig\.|table|tabla|box|cuadro)\b",
    ]

    cleaned: List[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        lower = line.lower()
        if any(re.match(pat, lower) for pat in header_patterns):
            continue
        if any(re.match(pat, lower) for pat in entry_patterns):
            continue
        cleaned.append(raw_line)
    return "\n".join(cleaned).strip()


def filter_toc_eval_text(text: str) -> str:
    if not text:
        return ""

    kept: List[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if re.match(r"^\d+\.\d+", line):
            continue
        if not re.search(r"\b\d+\b", line):
            continue
        kept.append(raw_line)
    return "\n".join(kept).strip()


def _extract_toc_lines(text: str, page_count: Optional[int]) -> str:
    """Extract TOC-like lines while preserving indentation."""
    lines = text.splitlines()
    output: List[str] = []
    pending_title: Optional[str] = None
    pending_indent = ""

    def flush_pending(page: str) -> None:
        nonlocal pending_title, pending_indent
        if not pending_title:
            return
        title = pending_title.strip()
        if page_count and page > page_count:
            pending_title = None
            pending_indent = ""
            return
        if page <= 0:
            pending_title = None
            pending_indent = ""
            return
        if _is_valid_toc_title(title):
            indent = pending_indent or _derive_indent(title)
            output.append(f"{indent}{title} ... {page}")
        pending_title = None
        pending_indent = ""

    for raw_line in lines:
        if not raw_line.strip():
            continue
        indent = raw_line[: len(raw_line) - len(raw_line.lstrip())]
        line = raw_line.rstrip()

        page_only = _page_only_value(line)
        if page_only is not None:
            flush_pending(page_only)
            continue

        match = _match_title_page(line)
        if match:
            title, page, has_dots, has_number_prefix = match
            if page_count and page > page_count:
                pending_title = None
                pending_indent = ""
                continue
            if page <= 0:
                pending_title = None
                pending_indent = ""
                continue
            if not (has_dots or has_number_prefix):
                pending_title = None
                pending_indent = ""
                continue
            if not has_number_prefix and not indent and len(title.split()) <= 2:
                pending_title = None
                pending_indent = ""
                continue
            if _is_valid_toc_title(title):
                line_indent = indent or _derive_indent(title)
                output.append(f"{line_indent}{title} ... {page}")
            pending_title = None
            pending_indent = ""
            continue

        if _should_skip_line(line):
            pending_title = None
            pending_indent = ""
            continue

        if pending_title:
            pending_title = f"{pending_title.rstrip()} {line.strip()}"
        else:
            pending_title = line.strip()
            pending_indent = indent

    return "\n".join(output).strip()


def _match_title_page(line: str) -> Optional[Tuple[str, int, bool, bool]]:
    stripped = line.strip()
    match = re.match(
        r"^(?P<title>.+?)(?P<dots>\s*[\.·•∙⋅…]{2,}\s*|\s+)(?P<page>\d{1,4})\s*$",
        stripped,
    )
    if not match:
        return None
    title = match.group("title").strip()
    has_dots = bool(re.search(r"[\.·•∙⋅…]{2,}", match.group("dots")))
    has_number_prefix = bool(re.match(r"^\d+(\.\d+)*\b", title))
    return title, int(match.group("page")), has_dots, has_number_prefix


def _count_toc_like_lines(text: str, page_count: Optional[int]) -> int:
    count = 0
    for raw_line in text.splitlines():
        match = _match_title_page(raw_line)
        if not match:
            continue
        title, page, has_dots, has_number_prefix = match
        if page_count and page > page_count:
            continue
        if page <= 0:
            continue
        if not (has_dots or has_number_prefix):
            continue
        if _is_valid_toc_title(title):
            count += 1
    return count


def _page_only_value(line: str) -> Optional[int]:
    match = re.match(r"^\s*(\d{1,4})\s*$", line)
    if not match:
        return None
    return int(match.group(1))


def _is_valid_toc_title(title: str) -> bool:
    letters = sum(ch.isalpha() for ch in title)
    if letters < 3:
        return False
    if re.match(r"^[\d\W_]+$", title):
        return False
    stripped = title.strip()
    if not stripped:
        return False
    first = stripped[0]
    if not (first.isupper() or first.isdigit()):
        return False
    if " | " in stripped:
        return False
    if re.search(r"\b(19|20)\d{2}\b", stripped) and not re.match(
        r"^\d+(\.\d+)*\b", stripped
    ):
        return False
    if len(stripped.split()) > 20 and not re.match(r"^\d+(\.\d+)*\b", stripped):
        return False
    if len(stripped) > 180:
        return False
    return True


def _derive_indent(title: str) -> str:
    match = re.match(r"^\s*(\d+(?:\.\d+)*)", title)
    if not match:
        return ""
    depth = match.group(1).count(".")
    return "  " * depth


def _should_skip_line(line: str) -> bool:
    lower = line.strip().lower()
    if re.match(r"^(table\s+of\s+contents|contents|sommaire|indice|índice)$", lower):
        return True
    if re.match(r"^[A-Z0-9/._-]{4,}$", line.strip()):
        return True
    if len(line.strip()) <= 2:
        return True
    return False


def _looks_like_toc_page(text: str) -> bool:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) < 5:
        return False
    toc_lines = 0
    for line in lines:
        if re.search(r"\.\s*\d+$", line) or re.search(r"\s\d+$", line):
            toc_lines += 1
    return toc_lines >= 3


def _format_toc_list(toc_data: List[Dict[str, Any]]) -> str:
    toc_lines: List[str] = []
    for entry in toc_data:
        page = entry.get("page", "?")
        level = entry.get("level", 1)
        label = entry.get("label") or entry.get("title") or entry.get("text") or ""
        toc_lines.append(f"[H{level}] {label} | page {page}")
    return "\n".join(toc_lines)


def _payload_has_toc(payload: dict) -> bool:
    return bool(
        payload.get("toc_classified")
        or payload.get("sys_toc_classified")
        or payload.get("toc")
        or payload.get("sys_toc")
    )


def _resolve_pdf_path(payload: dict) -> tuple[str, str]:
    raw_path = (
        payload.get("filepath")
        or payload.get("local_path")
        or payload.get("file_path")
        or payload.get("pdf_path")
        or payload.get("sys_filepath")
        or payload.get("sys_file_path")
        or payload.get("sys_local_path")
        or payload.get("sys_pdf_path")
    )
    if not raw_path:
        return "", "No PDF path fields found in payload"

    path = Path(raw_path)
    if path.is_absolute() and path.exists():
        return str(path), ""
    if path.is_absolute() and not path.exists():
        return "", f"PDF not found at {path}"

    app_root = Path(os.environ.get("APP_ROOT", PROJECT_ROOT))
    candidate = app_root / path
    if candidate.exists():
        return str(candidate), ""

    data_mount_path = os.environ.get("DATA_MOUNT_PATH")
    if data_mount_path:
        candidate = Path(data_mount_path) / path
        if candidate.exists():
            return str(candidate), ""

    tried = [str(app_root / path)]
    if data_mount_path:
        tried.append(str(Path(data_mount_path) / path))
    return "", f"PDF not found. Tried: {', '.join(tried)}"


def _truncate_text(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 200] + "\n...[truncated]...\n" + text[-150:]


def _build_prompt_content(template, pdf_extracted_toc: str, corrected_toc: str) -> str:
    pdf_excerpt = _truncate_text(pdf_extracted_toc, MAX_PDF_TOC_CHARS)
    prompt = template.render(
        pdf_extracted_toc=pdf_excerpt,
        corrected_toc=corrected_toc,
    )
    if len(prompt) <= MAX_RENDERED_PROMPT_CHARS:
        return prompt

    overflow = len(prompt) - MAX_RENDERED_PROMPT_CHARS
    new_len = max(1000, len(pdf_excerpt) - overflow - 200)
    pdf_excerpt = _truncate_text(pdf_excerpt, new_len)
    return template.render(
        pdf_extracted_toc=pdf_excerpt,
        corrected_toc=corrected_toc,
    )


def _format_db_toc_for_log(corrected_toc: str) -> str:
    lines = []
    for raw_line in corrected_toc.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match(r"^\[H(?P<level>\d+)\]\s*(?P<title>.+)$", line)
        if match:
            level = int(match.group("level"))
            title = match.group("title").strip()
            indent = "  " * max(level - 1, 0)
            lines.append(f"{indent}{title}")
        else:
            lines.append(line)
    return "\n".join(lines)


def _append_row(
    ws,
    row: List[str],
    pdf_link: Optional[str],
) -> None:
    ws.append(row)
    if pdf_link:
        cell = ws.cell(row=ws.max_row, column=len(row))
        cell.hyperlink = pdf_link
        cell.value = pdf_link


def _write_doc_log(
    doc_id: str,
    title: str,
    pdf_path: Optional[str],
    pdf_link: Optional[str],
    pdf_toc: str,
    db_toc: str,
    judge_result: str,
    judge_reason: str,
) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(LOG_PATH, "a", encoding="utf-8") as handle:
        handle.write("=" * 80 + "\n")
        handle.write(f"doc_id: {doc_id}\n")
        handle.write(f"title: {title}\n")
        if pdf_path:
            handle.write(f"pdf_path: {pdf_path}\n")
        if pdf_link:
            handle.write(f"pdf_link: {pdf_link}\n")
        handle.write("\nPDF TOC (extracted):\n")
        handle.write((pdf_toc.strip() or "(missing)") + "\n")
        handle.write("\nDB TOC (indented):\n")
        db_toc_rendered = _format_db_toc_for_log(db_toc).strip()
        handle.write((db_toc_rendered or "(missing)") + "\n\n")
        handle.write("LLM Judge:\n")
        handle.write(f"result: {judge_result}\n")
        handle.write(f"reason: {judge_reason}\n\n")


def sanitize_for_excel(text: str) -> str:
    if not text:
        return ""
    sanitized = ""
    for char in text:
        code = ord(char)
        if code >= 32 or code in (9, 10, 13):
            sanitized += char
        else:
            sanitized += " "
    sanitized = "".join(
        (
            char
            if unicodedata.category(char)[0] != "C" or char in ("\t", "\n", "\r")
            else " "
        )
        for char in sanitized
    )
    return sanitized


def _resolve_llm_config(model_hint: str) -> Tuple[Dict[str, Any], str]:
    config = load_datasources_config()
    supported_llms = config.get("supported_llms", {})
    if model_hint in supported_llms:
        return supported_llms[model_hint], model_hint
    for key, cfg in supported_llms.items():
        if cfg.get("model") == model_hint:
            return cfg, key
    raise ValueError(f"{model_hint} not found in supported_llms config")


def evaluate_toc_with_llm(pdf_extracted_toc: str, corrected_toc: str) -> dict:
    eval_model_config, model_key = _resolve_llm_config(JUDGE_LLM)

    llm_model = eval_model_config.get("model")
    llm_provider = eval_model_config.get("provider", "huggingface")
    llm_inference_provider = eval_model_config.get("inference_provider")

    if not llm_model:
        raise ValueError(f"Model not specified for {model_key} in config")

    llm = get_llm(
        provider=llm_provider,
        model=llm_model,
        temperature=0.0,
        max_tokens=512,
        inference_provider=llm_inference_provider,
    )

    validation_tmpl = jinja_env.get_template("toc_validation.j2")
    prompt_content = _build_prompt_content(
        validation_tmpl, pdf_extracted_toc, corrected_toc
    )

    response = llm.invoke([HumanMessage(content=prompt_content)])
    response_text = str(response.content).strip()

    json_match = re.search(r"```json\s*(\{.*?\})\s*```", response_text, re.DOTALL)
    if json_match:
        response_text = json_match.group(1)
    else:
        json_match = re.search(r"\{.*\}", response_text, re.DOTALL)
        if json_match:
            response_text = json_match.group(0)

    try:
        eval_result = json.loads(response_text)
        if "result" not in eval_result or "reason" not in eval_result:
            raise ValueError("Missing required fields in LLM response")
        if eval_result["result"] not in ["yes", "no"]:
            raise ValueError(f"Invalid result value: {eval_result['result']}")
        eval_result["rendered_prompt"] = prompt_content
        return eval_result
    except (json.JSONDecodeError, ValueError) as exc:
        return {
            "result": "error",
            "reason": (
                f"Failed to parse LLM response: {str(exc)}. "
                f"Raw response: {response_text[:200]}"
            ),
            "rendered_prompt": prompt_content,
        }


def _create_workbook() -> (
    Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TOC Evaluation"

    headers = [
        "doc_id",
        "title",
        "data_source",
        "eval_result",
        "eval_reason",
        "rendered_prompt",
        "pdf_link",
    ]
    ws.append(headers)

    header_fill = PatternFill(
        start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
    )
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    return wb, ws


def main():
    args = parse_args()

    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    LOG_PATH.write_text("", encoding="utf-8")

    if args.file_id:
        print(f"Fetching document by ID: {args.file_id}")
        docs, detected_data_source = get_document_by_id(args.file_id)
        data_source = detected_data_source
        if args.data_source != "uneg" and args.data_source != detected_data_source:
            print(
                f"Warning: Document found in '{detected_data_source}' collection, "
                f"but --data-source was set to '{args.data_source}'. "
                f"Using '{detected_data_source}'."
            )
    else:
        docs = get_test_documents(args.records, data_source=args.data_source)
        data_source = args.data_source

    print("Mode: Judge DB TOCs\n")

    wb, ws = _create_workbook()
    total = 0
    counts = {"yes": 0, "no": 0, "error": 0}

    for i, doc in enumerate(docs):
        doc_id = str(doc.id)
        total += 1
        payload = doc.payload
        print(f"Processing ({i+1}/{len(docs)}): {doc_id}")

        doc_title = get_document_title(payload)
        corrected_toc = get_corrected_toc_from_payload(payload)
        pdf_path, _ = _resolve_pdf_path(payload)
        pdf_link = get_pdf_web_link(payload)
        pdf_extracted_toc, pdf_error = extract_pdf_toc_for_validation(payload)
        if not corrected_toc:
            counts["error"] += 1
            _write_doc_log(
                doc_id,
                doc_title,
                pdf_path,
                pdf_link,
                pdf_extracted_toc,
                "",
                "error",
                "No TOC found in database",
            )
            _append_row(
                ws,
                [
                    doc_id,
                    doc_title,
                    data_source,
                    "error",
                    "No TOC found in database",
                    "",
                    pdf_link or "",
                ],
                pdf_link,
            )
            continue

        if not pdf_extracted_toc:
            counts["error"] += 1
            _write_doc_log(
                doc_id,
                doc_title,
                pdf_path,
                pdf_link,
                "",
                corrected_toc,
                "error",
                pdf_error or "Could not extract TOC from PDF",
            )
            _append_row(
                ws,
                [
                    doc_id,
                    doc_title,
                    data_source,
                    "error",
                    pdf_error or "Could not extract TOC from PDF",
                    "",
                    pdf_link or "",
                ],
                pdf_link,
            )
            continue

        eval_result = evaluate_toc_with_llm(pdf_extracted_toc, corrected_toc)
        result_value = eval_result.get("result", "error")
        judge_reason = sanitize_for_excel(eval_result.get("reason", ""))
        counts[result_value if result_value in counts else "error"] += 1
        _write_doc_log(
            doc_id,
            doc_title,
            pdf_path,
            pdf_link,
            pdf_extracted_toc,
            corrected_toc,
            result_value,
            judge_reason,
        )
        _append_row(
            ws,
            [
                doc_id,
                doc_title,
                data_source,
                result_value,
                judge_reason,
                sanitize_for_excel(eval_result.get("rendered_prompt", "")),
                pdf_link or "",
            ],
            pdf_link,
        )

    wb.save(args.output)
    print(f"Saved results to {args.output}")
    print(
        "Summary: total={total}, yes={yes}, no={no}, error={error}".format(
            total=total, yes=counts["yes"], no=counts["no"], error=counts["error"]
        )
    )


if __name__ == "__main__":
    main()
